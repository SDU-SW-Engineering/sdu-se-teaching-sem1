#!/usr/bin/env python3

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
from functools import cmp_to_key
import json

src_sheet_names = [
  "Software Engineering",
  "Software teknologi",
  "Spiludvikling og Læringsteknolo",
]

eduname2line = {
  "Software Engineering":            "Software Engineering",
  "Software teknologi":              "Softwareteknologi",
  "Spiludvikling og Læringsteknolo": "Spiludvikling og Læringsteknologi",
}

oop_dates = [
  "Mandag den 16. Januar, 2023",
  "Tirsdag den 17. Januar, 2023",
  "Onsdag den 18. Januar, 2023",
  "Torsdag den 19. Januar, 2023",
  "Fredag den 20. Januar, 2023",
]

oop_line_order = [
  "Spiludvikling og Læringsteknologi",
  "Software Engineering",
  "Softwareteknologi",
]

header = {
  # Tid
  "time": {
    "title": "Tid",
    "index": 0,
    "major": True,
    "width": 3,
    "color": "time",
  },
  "time.meet": {
    "title": "Møde",
    "index": 0,
    "major": False,
    "colwidth": 7,
    "color": "time",
    "comment": "Mødetid er op til 1 time før forventet starttid for effektivt at kunne håndtere udeblivelser",
  },
  "time.from": {
    "title": "Start",
    "index": 1,
    "major": False,
    "colwidth": 7,
    "color": "time",
  },
  "time.to": {
    "title": "Afslutning",
    "index": 2,
    "major": False,
    "colwidth": 10,
    "color": "time",
  },
  
  # Studerende
  "studerede": {
    "title": "Studerende",
    "index": 3,
    "major": True,
    "width": 3,
    "color": "student",
  },
  "name": {
    "title": "Navn",
    "index": 3,
    "major": False,
    "colwidth": 35,
    "color": "student",
  },
  "email": {
    "title": "Email",
    "index": 4,
    "major": False,
    "colwidth": 25,
    "color": "student",
  },
  "edu": {
    "title": "Retning",
    "index": 5,
    "major": False,
    "colwidth": 20,
    "color": "student",
  },
  
  # TA
  "ta": {
    "title": "Pointgivende Aktiviteter",
    "index": 6,
    "major": True,
    "width": 4,
    "grade": "noshow",
    "color": "pa",
  },
  "ta1": {
    "title": "PA1",
    "index": 6,
    "major": False,
    "colwidth": 6,
    "grade": "noshow",
    "color": "pa",
  },
  "ta2": {
    "title": "PA2",
    "index": 7,
    "major": False,
    "colwidth": 6,
    "grade": "noshow",
    "color": "pa",
  },
  "ta3": {
    "title": "PA3",
    "index": 8,
    "major": False,
    "colwidth": 6,
    "grade": "noshow",
    "color": "pa",
  },
  "ta.sum": {
    "title": "Avg",
    "index": 9,
    "major": False,
    "colwidth": 6,
    "grade": "noshow",
    "color": "pa",
  },
  
  # oral
  "oral": {
    "title": "Mundtlig Eksamen",
    "index": 10,
    "major": True,
    "width": 3,
    "grade": "noshow",
    "color": "oral",
  },
  "topic": {
    "title": "Emne",
    "index": 10,
    "major": False,
    "grade": "noshow",
    "color": "oral",
    "comment": "Hvilket emne blev trukket til præsentation?",
  },
  "exercise": {
    "title": "Opgave",
    "index": 11,
    "major": False,
    "grade": "noshow",
    "color": "oral",
    "comment": "Hvilken praktisk opgave blev trukket til løsning på tavle?",
  },
  "grade.oral": {
    "title": "Karakter",
    "index": 12,
    "major": False,
    "grade": "noshow",
    "color": "oral",
    "comment": "På skala fra 0 til 100",
  },
  
  # adjusted grade
  "adj.major": {
    "title": "Justeret",
    "index": 13,
    "major": True,
    "grade": "noshow",
    "color": "adjusted",
  },
  "grade.adj": {
    "title": "Karakter",
    "index": 13,
    "major": False,
    "grade": "noshow",
    "color": "adjusted",
  },
  
  # final grade
  "final.major": {
    "title": "Endelig",
    "index": 14,
    "major": True,
    "grade": "noshow",
    "color": "final",
  },
  "grade.final": {
    "title": "Karakter",
    "index": 14,
    "major": False,
    "grade": "noshow",
    "color": "final",
  },
}

color = {
  "time":     "FDDBFA",
  "student":  "F1F1F1",
  "pa":       "FCFDDB",
  "oral":     "DBFCFD",
  "adjusted": "FDDFDB",
  "final":    "DBFDE3",
  "table":    "F4DBFD",
}

group2size = {}
classes = ["1", "2", "3", "4"]

with open("sem_censors.json") as fo:
  sem_censors = json.loads("".join(fo.readlines()))
  
with open("sem_advisors.json") as fo:
  sem_advisors = json.loads("".join(fo.readlines()))

with open("student_mapping_override.json") as fo:
  student_mapping_override = json.loads("".join(fo.readlines()))

with open("oop_censors.json") as fo:
  oop_censors = json.loads("".join(fo.readlines()))
  
# 1-2: se, 3-4: st
sem_groups = [
  {"group": "1.5", "edu": "se", "day": "Jan 23", "from": "9:00", "to": "10:20"},
  {"group": "1.4", "edu": "se", "day": "Jan 23", "from": "10:40", "to": "12:20"},
  {"group": "2.2", "edu": "se", "day": "Jan 23", "from": "13:20", "to": "15:00"},
  {"group": "2.4", "edu": "se", "day": "Jan 23", "from": "15:20", "to": "17:00"},
  {"group": "1.1", "edu": "se", "day": "Jan 24", "from": "9:00", "to": "11:00"},
  {"group": "1.2", "edu": "se", "day": "Jan 24", "from": "11:20", "to": "13:00"},
  {"group": "1.3", "edu": "se", "day": "Jan 24", "from": "14:00", "to": "16:00"},
  {"group": "2.1", "edu": "se", "day": "Jan 26", "from": "9:00", "to": "10:40"},
  {"group": "2.5", "edu": "se", "day": "Jan 26", "from": "11:00", "to": "13:00"},
  {"group": "2.3", "edu": "se", "day": "Jan 26", "from": "14:00", "to": "16:00"},
  {"group": "3.2", "edu": "st", "day": "Jan 24", "from": "9:00", "to": "10:40"},
  {"group": "3.4", "edu": "st", "day": "Jan 24", "from": "11:00", "to": "12:40"},
  {"group": "3.1", "edu": "st", "day": "Jan 24", "from": "13:40", "to": "15:40"},
  {"group": "4.4", "edu": "st", "day": "Jan 25", "from": "9:00", "to": "10:40"}, # not Jan 26
  {"group": "4.1", "edu": "st", "day": "Jan 25", "from": "11:00", "to": "12:40"},
  {"group": "4.2", "edu": "st", "day": "Jan 25", "from": "13:40", "to": "15:40"},
  {"group": "4.3", "edu": "st", "day": "Jan 26", "from": "9:00", "to": "10:40"},
  {"group": "3.3", "edu": "st", "day": "Jan 26", "from": "11:00", "to": "12:40"},
  {"group": "3.5", "edu": "st", "day": "Jan 26", "from": "13:40", "to": "15:20"},
]

oop_slots = [
  {"from": "9:00", "to": "9:20", "meet": "9:00"},
  {"from": "9:20", "to": "9:40", "meet": "9:00"},
  {"from": "9:40", "to": "10:00", "meet": "9:00"},
  {"from": "10:00", "to": "10:20", "meet": "9:00"},
  {"from": "10:20", "to": "10:40", "meet": "9:20"},
  {"from": "10:40", "to": "11:00", "meet": "9:40"},
  {"from": "11:00", "to": "11:20", "meet": "10:00"},
  {"from": "11:20", "to": "11:40", "meet": "10:20"},
  {"from": "11:40", "to": "12:00", "meet": "10:40"},
  {"from": "12:00", "to": "12:20", "meet": "11:00"},
  {"from": "12:20", "to": "12:40", "meet": "11:20", "break": "lunch"},
  {"from": "12:40", "to": "13:00", "meet": "11:40", "break": "skip"},
  {"from": "13:00", "to": "13:20", "meet": "12:00", "break": "skip"},
  {"from": "13:20", "to": "13:40", "meet": "12:20"},
  {"from": "13:40", "to": "14:00", "meet": "12:40"},
  {"from": "14:00", "to": "14:20", "meet": "13:00"},
  {"from": "14:20", "to": "14:40", "meet": "13:20"},
  {"from": "14:40", "to": "15:00", "meet": "13:40"},
  {"from": "15:00", "to": "15:20", "meet": "14:00"},
  {"from": "15:20", "to": "15:40", "meet": "14:20"},
  {"from": "15:40", "to": "16:00", "meet": "14:40", "break": "break"},
  {"from": "16:00", "to": "16:20", "meet": "15:00"},
  {"from": "16:20", "to": "16:40", "meet": "15:20"},
  {"from": "16:40", "to": "17:00", "meet": "15:40"},
  {"from": "17:00", "to": "17:20", "meet": "16:00"},
  {"from": "17:20", "to": "17:40", "meet": "16:20"},
  {"from": "17:40", "to": "18:00", "meet": "16:40"},
  {"from": "18:00", "to": "18:20", "meet": "17:00"},
  {"from": "18:20", "to": "18:40", "meet": "17:20"},
  {"from": "18:40", "to": "19:00", "meet": "17:40"},
]

def groups2thold (groups):
  for group in groups:
    if len(group)==2 and group[0]=="T":
      return group
  return -1

def groups2group (groups):
  for group in groups:
    if group.startswith("Gruppe "):
      return group
  return -1

def load_datafile (filename, students=None):
  if students==None: students = []
  
  with open(filename) as fo:
    lines = fo.readlines()
  
  for line in lines:
    elements = line.split("\t")
    name   = elements[1].strip()
    email  = elements[2].strip()
    role   = elements[3].strip()
    groups = elements[4].split(", ")
    thold  = groups2thold(groups)
    group  = groups2group(groups)
    
    if not role in ["Studerende", "Student"]: continue
    
    student = {
      "name":  name,
      "email": email,
      "role":  role,
      "thold": thold,
      "group": group,
    }
    students.append(student)
  
  return students

def load_student_lines (filename):
  global name2line
  
  wb = load_workbook(filename=filename)
  
  for sheet_name in src_sheet_names:
    line = eduname2line[sheet_name]
    sheet = wb[sheet_name]
    
    for row in range(2, 200):
      if sheet["A%d"%row].value==None:
        
        break
      
      gname = sheet["B%d"%row].value
      fname = sheet["C%d"%row].value
      name = "%s %s" % (gname, fname)
      
      name2line[name] = line
  
  # override
  for name in student_mapping_override:
    name2line[name] = student_mapping_override[name]

def load_group_sizes (students):
  global group2size
  
  for student in students:
    group = student["group"]
    
    if not group in group2size:
      group2size[group] = 0
    
    group2size[group] += 1

def load_oop_ta12_scores (filename, students, key):
  wb = load_workbook(filename=filename)
  sheet = wb["QuestionResults"]
  
  for row in range(2, 200):
    if sheet["A%d"%row].value==None:
      break
    
    score    = str(sheet["E%d"%row].value)
    email    = sheet["B%d"%row].value.strip()
    
    for student in students:
      if student["email"]==email:
#        print("%s: %s -> %s" % (email, student["name"], score))
        student[key]=score
        break

def load_oop_ta3_scores (filename, students):
  with open(filename) as fo:
    lines = fo.readlines()
#  wb = load_workbook(filename=filename)
#  sheet = wb["TA3 Scores"]
  
  for line in lines[1:]:
#  for row in range(2, 200):
#    if sheet["A%d"%row].value==None:
#      break
    
    cols = line.strip().split(",")
    username = cols[0]
    score    = str(float(cols[11])*100)
#    username = sheet["A%d"%row].value
#    score    = sheet["L%d"%row].value[:-1]
    email    = "%s@student.sdu.dk" % username
    
    for student in students:
      if student["email"]==email:
#        print("%s: %s -> %s" % (email, student["name"], score))
        student["ta3"]=score
        break

def import_groups (oop_students, sem_students):
  name2group = {}
  
  for entry in sem_students:
    name2group[entry['name']] = entry["group"]
  
  for entry in oop_students:
    name = entry["name"]
    entry["group"] = name2group[name] if name in name2group else "-1"

def generate_line2index ():
  line2index = {}
  
  for i in range(len(oop_line_order)):
    line = oop_line_order[i]
    line2index[line] = i
  
  return line2index

def sort_students (students):
  def compare (s1, s2):
    def strcmp (s1, s2):
      return -1 if s1<s2 else (0 if s1==s2 else 1)
    
    n1 = s1["name"]
    n2 = s2["name"]
    
    # primary
    l1 = name2line[n1] if n1 in name2line else ""
    l2 = name2line[n2] if n2 in name2line else ""
    if l1 != l2:
      return strcmp(line2index[l1], line2index[l2])
    
    # secondary
    if s1["group"] != s2["group"]:
      return strcmp(s1["group"], s2["group"])
    
    # tertiary
    return strcmp(n1, n2)
  
  students.sort(key=cmp_to_key(compare))

def split_students_on_educations (students):
  structure = {}
  
  for student in students:
#    print(name2line.keys())
    if not student["name"] in name2line:
      print("'"+student["name"]+"'", type(student["name"]))
      continue
    line = name2line[student["name"]]
    if not line in structure: structure[line] = []
    structure[line].append(student)
  
  return structure

def x2col (x):
  return "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[x]

def xy2cell (x, y): # zero-indexed
  col = x2col(x)
  return "%s%d" % (col, y+1)

def insert_students (sheet, students, show_grades):
  row = 5
  sloti = 0
  for student in students:
    # init
    slot = oop_slots[sloti]
    
    # handle break
    while "break" in slot:
      if slot["break"]!="skip":
        cell = xy2cell(0, row)
        sheet[cell].font = Font(i=True)
        sheet[cell].value = "pause"
        row += 1
      sloti += 1
      slot = oop_slots[sloti]
    
    # insert
#    print(sheet[xy2cell(0, row)].value)
#    print(student)
    for i in range(0,3):
      sheet[xy2cell(i, row)].fill = PatternFill("solid", fgColor=color["time"])
    sheet[xy2cell(0, row)].value = slot["meet"]
    sheet[xy2cell(1, row)].value = slot["from"]
    sheet[xy2cell(2, row)].value = slot["to"]
    for i in range(3,6):
      sheet[xy2cell(i, row)].fill = PatternFill("solid", fgColor=color["student"])
    sheet[xy2cell(3, row)].value = student["name"]
    sheet[xy2cell(4, row)].value = student["email"]
    sheet[xy2cell(5, row)].value = name2line[student["name"]]
    if show_grades:
      for i in range(6,10):
        sheet[xy2cell(i, row)].fill = PatternFill("solid", fgColor=color["pa"])
      sheet[xy2cell(6, row)].value = student["ta1"] if "ta1" in student else "0"
      sheet[xy2cell(7, row)].value = student["ta2"] if "ta2" in student else "0"
      sheet[xy2cell(8, row)].value = student["ta3"] if "ta3" in student else "0"
      sheet[xy2cell(9, row)].value = "=(%s+%s+%s)/3" % (xy2cell(6, row), xy2cell(7, row), xy2cell(8, row))
      for i in range(10,13):
        sheet[xy2cell(i, row)].fill = PatternFill("solid", fgColor=color["oral"])
      for i in range(13,14):
        sheet[xy2cell(i, row)].fill = PatternFill("solid", fgColor=color["adjusted"])
      sheet[xy2cell(13, row)].value = "=%s+(%s/10)" % (xy2cell(12, row), xy2cell(9, row))
      for i in range(14,15):
        sheet[xy2cell(i, row)].fill = PatternFill("solid", fgColor=color["final"])
      formula = "=if(N%d>T6,12,if(N%d>T7,10,if(N%d>T8,7,if(N%d>T9,4,if(N%d>T10,2,if(N%d>T11,0,-3))))))"
      sheet[xy2cell(14, row)].value = formula % (row+1, row+1, row+1, row+1, row+1, row+1)
    
    # update
    row += 1
    sloti += 1

def generate_oop_schedules (filename, show_censors, show_grades):
  wb = Workbook()
  
  # produce sheets
  sheets = {"Aslak": {}, "Peter": {}}
  for date in oop_dates:
    day = date.split(" ")[0]
    for examiner in ["Aslak", "Peter"]:
      # guard
      if examiner=="Peter" and day=="Tirsdag": continue
      
      sheet_title = "%s %s" % (day, examiner)
      wb.create_sheet(title = sheet_title)
      sheet = wb[sheet_title]
      
      # title
      sheet["A1"].font = Font(b=True)
      sheet["A1"].value = "%s/%s" % (date, examiner)
      
      # censor
      if show_censors:
        sheet["A2"].font = Font(i=True)
        sheet["A2"].value = "Censor: %s (%s)" % (oop_censors[examiner][day]["name"], oop_censors[examiner][day]["email"])
      
      # header
      for key in header:
        entry = header[key]
        cell = xy2cell(entry["index"], 3 if entry["major"] else 4)
        
        # guard
        if "grade" in entry and not show_grades: continue
        
        if "width" in entry:
          sheet.merge_cells(start_row=1+(3 if entry["major"] else 4), \
                            start_column=1+entry["index"], \
                            end_row=1+(3 if entry["major"] else 4), \
                            end_column=1+entry["index"]+entry["width"]-1)
        sheet[cell].font = Font(b=True)
        sheet[cell].value = entry["title"]
        
        if "comment" in entry:
          sheet[cell].comment = Comment(entry["comment"], "Aslak Johansen")
        
        if "color" in entry:
          sheet[cell].fill = PatternFill("solid", fgColor=color[entry["color"]])
        
        if "colwidth" in entry:
          sheet.column_dimensions[x2col(entry["index"])].width = entry["colwidth"]
      
      # conversion table
      sheet.column_dimensions[x2col(15)].width = 5
      sheet.column_dimensions[x2col(16)].width = 9
      sheet.column_dimensions[x2col(17)].width = 10
      sheet.column_dimensions[x2col(18)].width = 12
      sheet.column_dimensions[x2col(19)].width = 11
      for i, value in [(0,"Karakter"), (1,"Point/min"), (2,"Point/mean"), (3,"Point/max")]:
        cell = xy2cell(16+i, 4)
        sheet[cell].font = Font(b=True)
        sheet[cell].fill = PatternFill("solid", fgColor=color["table"])
        sheet[cell].value = value
      for i, grade, pmin, pmax in [(0, "12", 92, 100), (1,"10", 81, 91), (2,"7", 66, 80), (3,"4", 56, 65), (4, "2", 50, 55), (5, "00", 16, 49), (6, "-3", 0, 15)]:
        cell0 = xy2cell(16+0, 5+i)
        cell1 = xy2cell(16+1, 5+i)
        cell2 = xy2cell(16+2, 5+i)
        cell3 = xy2cell(16+3, 5+i)
        sheet[cell0].fill = PatternFill("solid", fgColor=color["table"])
        sheet[cell0].value = grade
        sheet[cell1].fill = PatternFill("solid", fgColor=color["table"])
        sheet[cell1].value = pmin
        sheet[cell2].fill = PatternFill("solid", fgColor=color["table"])
        sheet[cell2].value = "=average(%s,%s)" % (cell1, cell3)
        sheet[cell3].fill = PatternFill("solid", fgColor=color["table"])
        sheet[cell3].value = pmax
      
      # register sheet
      sheets[examiner][day] = sheet
  
  # data: split input
  students = split_students_on_educations(oop_students)
  
  # data: Spiludvikling og Læringsteknologi || Software Engineering
  if True:
    gamer = students["Spiludvikling og Læringsteknologi"]
    sweng = students["Software Engineering"]
    
    # sanity check: Software Engineering
    sanity = list(filter(lambda e: not e["thold"] in ["T1", "T2"], sweng))
    if len(sanity)>0:
      print("ERR: Sanity check for generate_oop_schedules/data/Software Engineering failed:")
      for entry in sanity:
        thold = "T2"
        print(" - %s, assigning %s" % (entry, thold))
        entry["thold"] = thold
    
    # sanity check: Spiludvikling og Læringsteknologi
    sanity = list(filter(lambda e: not e["thold"] in ["T5", "T6"], gamer))
    if len(sanity)>0:
      print("ERR: Sanity check for generate_oop_schedules/data/Spiludvikling og Læringsteknologi failed:")
      for entry in sanity:
        print(" - %s" % entry)
    
    t1 = list(filter(lambda e: e["thold"]=="T1", sweng)) # aslak
    t2 = list(filter(lambda e: e["thold"]=="T2", sweng)) # peter
    t5 = list(filter(lambda e: e["thold"]=="T5", gamer)) # aslak
    t6 = list(filter(lambda e: e["thold"]=="T6", gamer)) # peter
    
    print("OOP EXAM: Students in T1: %d" % len(t1))
    print("OOP EXAM: Students in T2: %d" % len(t2))
    print("OOP EXAM: Students in SWENG: %d" % (len(t1)+len(t2)))
    print("OOP EXAM: Students in T5: %d" % len(t5))
    print("OOP EXAM: Students in T6: %d" % len(t6))
    print("OOP EXAM: Students in GAMER: %d" % (len(t5)+len(t6)))
    
    teamaslak = t5 + t1
    teampeter = t6 + t2
    
    aslaksplit1 = int(len(teamaslak)/3)
    aslaksplit2 = int(2*len(teamaslak)/3)
    insert_students(sheets["Aslak"]["Mandag"] , teamaslak[:aslaksplit1], show_grades)
    insert_students(sheets["Aslak"]["Tirsdag"], teamaslak[aslaksplit1:aslaksplit2],show_grades)
    insert_students(sheets["Aslak"]["Onsdag"] , teamaslak[aslaksplit2:], show_grades)
    
    petersplit = int(len(teampeter)/2)
    insert_students(sheets["Peter"]["Mandag"] , teampeter[:petersplit], show_grades)
    insert_students(sheets["Peter"]["Onsdag"] , teampeter[petersplit:], show_grades)
  
  # data: Softwareteknologi
  if True:
    swtech = students["Softwareteknologi"]
    
    # sanity check
    sanity = list(filter(lambda e: not e["thold"] in ["T3", "T4"], swtech))
    if len(sanity)>0:
      print("ERR: Sanity check for generate_oop_schedules/data/Softwareteknologi failed:")
      for entry in sanity:
        print(" - %s" % entry)
    
    t3 = list(filter(lambda e: e["thold"]=="T3", swtech)) # aslak
    t4 = list(filter(lambda e: e["thold"]=="T4", swtech)) # peter
    
    print("OOP EXAM: Students in T3: %d" % len(t3))
    print("OOP EXAM: Students in T4: %d" % len(t4))
    print("OOP EXAM: Students in SWTEK: %d" % (len(t3)+len(t4)))
    
    t3split = int(len(t3)/2)
    insert_students(sheets["Aslak"]["Torsdag"], t3[:t3split], show_grades)
    insert_students(sheets["Aslak"]["Fredag"], t3[t3split:], show_grades)
    
    t4split = len(t4)-8
    insert_students(sheets["Peter"]["Torsdag"], t4[:t4split], show_grades)
    insert_students(sheets["Peter"]["Fredag"], t4[t4split:], show_grades)
  
  # remove original sheet
  wb.remove(wb['Sheet'])
  
  # save resulting workbook
  wb.save(filename)

def generate_sem_schedules (filename, show_censors):
  # tex file creation
  texlines = []
  texlines.append("\\documentclass{article}")
  texlines.append("\\usepackage[utf8]{inputenc}")
  texlines.append("\\title{Software Educations 1st Semester Project Exam 2022}")
  texlines.append("\\date{}")
  texlines.append("\\begin{document}")
  texlines.append("\\maketitle")
  
  texlines.append("\\section{Software Engineering}")
  for day in ["Jan 23", "Jan 24", "Jan 26"]:
    texlines.append("\\subsection{%s}" % day)
    texlines.append("\\textbf{Lokale:} %s" % ("U145" if day=="Jan 23" else "U143"))
    
    # censor
    if show_censors:
      censor = sem_censors["se"][day]
      texlines.append("\\\\")
      texlines.append("\\textbf{Censor:} %s (\\texttt{%s})%s" % (censor["name"], censor["email"], " [%s]"%censor["note"] if "note" in censor and censor["note"]!="" else ""))
    
    for group in sem_groups:
      if group["edu"]!="se": continue
      if group["day"]!= day: continue
      
      texlines.append("\\subsubsection{%s $\\rightarrow$ %s: Gruppe %s}" % (group["from"], group["to"], group["group"]))
      
      advisor = sem_advisors[group["group"]]
      texlines.append("\\textbf{Vejleder:} %s (\\texttt{%s})" % (advisor["name"], advisor["email"]))
      
      texlines.append("\\begin{itemize}")
      for student in sem_students:
        if student["group"]!="Gruppe %s"%group["group"]: continue
        texlines.append("  \\item %s (\\texttt{%s})" % (student["name"], student["email"]))
      texlines.append("\\end{itemize}")
  
  texlines.append("\\section{Software Teknologi}")
  texlines.append("\\textbf{Lokale:} U145")
  for day in ["Jan 24", "Jan 25", "Jan 26"]:
    texlines.append("\\subsection{%s}" % day)
    
    # censor
    if show_censors:
      censor = sem_censors["st"][day]
      texlines.append("\\textbf{Censor:} %s (\\texttt{%s})%s" % (censor["name"], censor["email"], " [%s]"%censor["note"] if "note" in censor and censor["note"]!="" else ""))
    
    for group in sem_groups:
      if group["edu"]!="st": continue
      if group["day"]!= day: continue
      
      texlines.append("\\subsubsection{%s $\\rightarrow$ %s: Gruppe %s}" % (group["from"], group["to"], group["group"]))
      
      advisor = sem_advisors[group["group"]]
      texlines.append("\\textbf{Vejleder:} %s (\\texttt{%s})" % (advisor["name"], advisor["email"]))
      
      texlines.append("\\begin{itemize}")
      for student in sem_students:
        if student["group"]!="Gruppe %s"%group["group"]: continue
        texlines.append("  \\item %s (\\texttt{%s})" % (student["name"], student["email"]))
      texlines.append("\\end{itemize}")
  
  texlines.append("\\end{document}")
  with open(filename, "w") as fo:
    fo.writelines(map(lambda line: "%s\n"%line, texlines))

def print_class_times ():
  for classname in classes:
    count = 0
    groups = []
    
    for student in sem_students:
      if student["thold"] == "T"+classname:
        count += 1
        if not student["group"] in groups:
          groups.append(student["group"])
    
    print("Class %s: %2d students in %d groups => %3d min / %1d h %s" % (classname, count, len(groups), 20*count, 20*count/60, list(map(lambda group: group2size[group]*20, groups))))

oop_students = []
load_datafile("oop1.data", oop_students)
load_datafile("oop2.data", oop_students)

sem_students = []
load_datafile("sem1.data", sem_students)
load_datafile("sem2.data", sem_students)

name2line = {}
load_student_lines("Lister SI1-OOP19 med klasser.xlsx")

load_oop_ta12_scores ("Pointgiven_Aktivitet_1_export_2023_01_11__01_23.xlsx", oop_students, "ta1")
load_oop_ta12_scores ("Pointgiven_Aktivitet_1_s_rlige_vilk_r__export_2023_01_11__01_24.xlsx", oop_students, "ta1")
load_oop_ta12_scores ("Pointgivende_Aktivitet_2_export_2023_01_11__01_21.xlsx", oop_students, "ta2")
load_oop_ta12_scores ("Pointgivende_Aktivitet_2_s_rlige_vilk_r__export_2023_01_11__01_23.xlsx", oop_students, "ta2")
#load_oop_ta3_scores("TA3 Scores.xlsx", oop_students)
load_oop_ta3_scores("TA3 Scores.csv", oop_students)

load_group_sizes(sem_students)

import_groups(oop_students, sem_students)
line2index = generate_line2index()
sort_students(oop_students)

generate_oop_schedules("SDU SEST 2022 OOP Exams.xlsx", show_censors=False, show_grades=False)
generate_oop_schedules("SDU SEST 2022 OOP Exams Full.xlsx", show_censors=True, show_grades=True)
generate_sem_schedules("SDU SEST 2022 Sem1 Project Exams.tex", show_censors=False)
generate_sem_schedules("SDU SEST 2022 Sem1 Project Exams Full.tex", show_censors=True)

#print(oop_students)
#print(sem_students)
#print(name2line)
#print(group2size)
#print_class_times()

