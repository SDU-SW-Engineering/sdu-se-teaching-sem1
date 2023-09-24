#!/usr/bin/env python3

from openpyxl import load_workbook, Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font

#input_filename = "Lister SI1-OOP19 med klasser.xlsx"
output_filename = "SDU SE OOP 2023 - Pointgivende Aktivitet %d.xlsx"

#sheet_names = [
#  "Software Engineering",
#  "Software teknologi",
#  "Spiludvikling og Læringsteknolo",
#]

classcount = 7
author = "Aslak Johansen"

syntax_guide = '''
Giv følgende bedømmelse:
0: Besvarelsen indeholder intet der minder om Javakode.
1: Besvarelsen indeholder spor af Javakode med mange og betydelige fejl.
2: Besvarelsen indeholder en del korrekt Javakode med nogle fejl og mangler.
3: Besvarelsen indeholder mest syntaktisk korrekt Javakode med få eller mest ubetydelige fejl.
4: Besvarelsen indeholder syntaktisk korrekt java.
'''

def load_studentlist (filename):
  names = []
  
  with open(filename) as fo:
    lines = fo.readlines()
    for line in lines:
      name = line.strip()
      if name=="": continue
      names.append(name)
  
  return names

def generate_sheets (pa, owb, sheeti, rowi):
  for i in range(classcount):
    name = 'Klasse %u'%(i+1)
    owb.create_sheet(title = name)
    sheet = owb[name]
    
    sheet["A3"].font = Font(b=True)
    sheet["A3"].value = "Fornavn"
#    sheet["B3"].font = Font(b=True)
#    sheet["B3"].value = "Efternavn"
    sheet.merge_cells(range_string="B1:H1")
    sheet["B1"].font = Font(b=True)
    sheet["B1"].fill = PatternFill("solid", fgColor="AAFFAA")
    sheet["B1"].value = "Opgave 9"
    sheet["B2"].font = Font(b=True)
    sheet["B2"].fill = PatternFill("solid", fgColor="CCFF88")
    sheet["B3"].font = Font(b=True)
    sheet["B3"].fill = PatternFill("solid", fgColor="CCFF88")
    sheet["B3"].value = "Syntaks"
    sheet["B3"].comment = Comment(syntax_guide, author)
    sheet.merge_cells(range_string="C2:G2")
    sheet["C2"].font = Font(b=True)
    sheet["C2"].fill = PatternFill("solid", fgColor="88FFCC")
    sheet["C2"].value = "Besvarelse"
    sheet["C3"].font = Font(b=True)
    sheet["C3"].fill = PatternFill("solid", fgColor="88FFCC")
    c3value = ["", "for(int i;i<10;i++)", "class Car"]
    c3comment = ["", "Erklæring af en for-løkke med initialisering af en int variabel ved navn \"i\" til værdien 0 og en continuation condition der tester om denne værdi er mindre end 10 og et update-statement der tæller værdien op med én for hvert gennemløb.", "En erklæring af klassen Car."]
    sheet["C3"].value = c3value[pa]
    sheet["C3"].comment = Comment("%s\n\n0 eller 1" % c3comment[pa], author)
    sheet["D3"].font = Font(b=True)
    sheet["D3"].fill = PatternFill("solid", fgColor="88FFCC")
    d3value = ["", "print(i)", "string licenseNumber"]
    d3comment = ["", "Udskrivelse af værdien for variablen \"i\" i bodyen.", "En attribut ved navn \"licenseNumber\" af typen String."]
    sheet["D3"].value = d3value[pa]
    sheet["D3"].comment = Comment("%s\n\n0 eller 1" % d3comment[pa], author)
    sheet["E3"].font = Font(b=True)
    sheet["E3"].fill = PatternFill("solid", fgColor="88FFCC")
    e3value = ["", "int i=i", "(get&&set)LicenseNumber"]
    e3comment = ["", "Erklæring af en ny variabel af typen int i bodyen intialiseret til den værdi variablen \"i\" har.", "Accessor- og mutator metoder til licenseNumber."]
    sheet["E3"].value = e3value[pa]
    sheet["E3"].comment = Comment("%s\n\n0 eller 1" % e3comment[pa], author)
    sheet["F3"].font = Font(b=True)
    sheet["F3"].fill = PatternFill("solid", fgColor="88FFCC")
    f3value = ["", "print(++j)", "Car(String)"]
    f3comment = ["", "Tæller den nye variabel op med én og udskriver dens værdi.", "En Constructor med én parameter - licenseNumber - hvis værdi anvendes til at initialisere licenseNumber attributten, og som kalder constructoren i Vehicle via \"super\" med en int."]
    sheet["F3"].value = f3value[pa]
    sheet["F3"].comment = Comment("%s\n\n0 eller 1" % f3comment[pa], author)
    sheet["G3"].font = Font(b=True)
    sheet["G3"].fill = PatternFill("solid", fgColor="88FFCC")
    sheet["G3"].value = "Sum"
    sheet["H3"].font = Font(b=True)
    sheet["H2"].fill = PatternFill("solid", fgColor="AAFFAA")
    sheet["H3"].fill = PatternFill("solid", fgColor="AAFFAA")
    sheet["H3"].value = "Resultat"
    sheet.merge_cells(range_string="I1:O1")
    sheet["I1"].font = Font(b=True)
    sheet["I1"].fill = PatternFill("solid", fgColor="AAAAFF")
    sheet["I1"].value = "Opgave 10"
    sheet["I2"].font = Font(b=True)
    sheet["I2"].fill = PatternFill("solid", fgColor="CC88FF")
    sheet["I3"].font = Font(b=True)
    sheet["I3"].fill = PatternFill("solid", fgColor="CC88FF")
    sheet["I3"].value = "Syntaks"
    sheet["I3"].comment = Comment(syntax_guide, author)
    sheet.merge_cells(range_string="J2:N2")
    sheet["J2"].font = Font(b=True)
    sheet["J2"].fill = PatternFill("solid", fgColor="88CCFF")
    sheet["J2"].value = "Besvarelse"
    sheet["J3"].font = Font(b=True)
    sheet["J3"].fill = PatternFill("solid", fgColor="88CCFF")
    j3value = ["", "int getLargerNumber (int arg)", "class Person implements Printable"]
    j3comment = ["", "Erklæring af en metode ved navn \"getLargerNumber\", der returnerer en \"int\" og tager en \"int\" som argument.", "En erklæring af klassen Person som implementerer interface Printable."]
    sheet["J3"].value = j3value[pa]
    sheet["J3"].comment = Comment("%s\n\n0 eller 1" % j3comment[pa], author)
    sheet["K3"].font = Font(b=True)
    sheet["K3"].fill = PatternFill("solid", fgColor="88CCFF")
    k3value = ["", "print(arg)", "String name"]
    k3comment = ["", "Udskriver værdien af argumentet.", "En attribut ved navn \"name\" af typen \"String\"."]
    sheet["K3"].value = k3value[pa]
    sheet["K3"].comment = Comment("%s\n\n0 eller 1" % k3comment[pa], author)
    sheet["L3"].font = Font(b=True)
    sheet["L3"].fill = PatternFill("solid", fgColor="88CCFF")
    l3value = ["", "arg++", "encapsulate(name)"]
    l3comment = ["", "Øger værdien af argumentet med én.", "Indkapsling af attributten \"name\" (private modifier) og en accessor metode."]
    sheet["L3"].value = l3value[pa]
    sheet["L3"].comment = Comment("%s\n\n0 eller 1" % l3comment[pa], author)
    sheet["M3"].font = Font(b=True)
    sheet["M3"].fill = PatternFill("solid", fgColor="88CCFF")
    m3value = ["", "return(42)", "@Override print"]
    m3comment = ["", "Returnerer den nye værdi.", "Override af metoden \"print\" således at værdien af attributten name udskrives."]
    sheet["M3"].value = m3value[pa]
    sheet["M3"].comment = Comment("%s\n\n0 eller 1" % m3comment[pa], author)
    sheet["N3"].font = Font(b=True)
    sheet["N3"].fill = PatternFill("solid", fgColor="88CCFF")
    sheet["N3"].value = "Sum"
    sheet["O2"].font = Font(b=True)
    sheet["O2"].fill = PatternFill("solid", fgColor="AAAAFF")
    sheet["O3"].font = Font(b=True)
    sheet["O3"].fill = PatternFill("solid", fgColor="AAAAFF")
    sheet["O3"].value = "Resultat"
    
    sheeti.append(sheet)
  
  owb.remove(owb['Sheet'])

def generate_pa (pa):
  rowi   = [4]*classcount
  sheeti = []

#  iwb = load_workbook(filename=input_filename)
  owb = Workbook()
  
  # generate sheets
  generate_sheets(pa, owb, sheeti, rowi)
  
  for tnumber in range(1,classcount+1):
    studentlist = load_studentlist("t%u.txt"%tnumber)
    print(studentlist)
    osheet = sheeti[tnumber-1]
    orow=4
    for student in studentlist:
      name = student
    
#  for sheet_name in sheet_names:
#    sheet = iwb[sheet_name]
#    
#    for row in range(2, 200):
#      if sheet["A%d"%row].value==None:
#        break
#      
#      gname = sheet["B%d"%row].value
#      fname = sheet["C%d"%row].value
#      lname = sheet["C%d"%row].value
#      clsname = sheet["E%d"%row].value
#      grpname = sheet["F%d"%row].value
      i = tnumber
      
#      orow   = rowi[i]
#      rowi[i] += 1
#      orow = 5
      
      osheet["A%d"%orow].value = name
      osheet["B%d"%orow].fill = PatternFill("solid", fgColor="CCFF88")
      for letter in ["C", "D", "E", "F", "G"]:
        osheet["%s%d"%(letter, orow)].fill = PatternFill("solid", fgColor="88FFCC")
      osheet["G%d"%orow].value = "=sum(C%d:F%d)"%(orow, orow)
      osheet["H%d"%orow].fill = PatternFill("solid", fgColor="AAFFAA")
      osheet["H%d"%orow].value = "=ceiling(((C%u*25)*0.5 + (H%u*25)*0.5)*15/100, 1)"%(orow, orow)
      osheet["I%d"%orow].fill = PatternFill("solid", fgColor="CC88FF")
      for letter in ["J", "K", "L", "M", "N"]:
        osheet["%s%d"%(letter, orow)].fill = PatternFill("solid", fgColor="88CCFF")
      osheet["N%d"%orow].value = "=sum(K%d:N%d)"%(orow, orow)
      osheet["O%d"%orow].fill = PatternFill("solid", fgColor="AAAAFF")
      osheet["O%d"%orow].value = "=ceiling(((J%u*25)*0.5 + (O%u*25)*0.5)*15/100, 1)"%(orow, orow)
      orow += 1
  
  owb.save(output_filename % pa)

for i in range(1,3):
  generate_pa(i)

